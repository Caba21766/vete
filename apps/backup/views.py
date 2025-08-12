
import os
import subprocess
from django.conf import settings
from django.shortcuts import render, redirect
from django.utils.timezone import now
from django.http import FileResponse, HttpResponseNotFound, HttpResponse

# Directorio donde se guardarán los respaldos
RESPALDO_DIR = os.path.join(settings.MEDIA_ROOT, 'respaldos')
os.makedirs(RESPALDO_DIR, exist_ok=True)

def crear_respaldo(request):
    nombre_archivo = f"respaldo_{now().strftime('%Y%m%d_%H%M%S')}.sql"
    ruta_respaldo = os.path.join(RESPALDO_DIR, nombre_archivo)

    ruta_mysql_dump = "mysqldump"  # Linux: se espera que esté en el PATH

    comando = [ruta_mysql_dump, "-u", "root", "-pRoot", "vete"]

    try:
        with open(ruta_respaldo, "w") as salida:
            resultado = subprocess.run(comando, stdout=salida, stderr=subprocess.PIPE, text=True)

        if resultado.returncode == 0:
            return redirect('backup:listar_respaldos')

        errores_limpios = "\n".join([
            linea for linea in resultado.stderr.splitlines()
            if "Advertencia" not in linea
        ])

        if errores_limpios.strip() == "":
            return redirect('backup:listar_respaldos')

        return HttpResponse(f"<h3>Error al crear el respaldo.</h3><pre>{errores_limpios}</pre>")

    except Exception as e:
        return HttpResponse(f"<h3>Error inesperado:</h3><pre>{str(e)}</pre>")


#------------------------------------
import shutil
import tempfile

def descargar_respaldo(request, nombre):
    """Permite descargar un respaldo SIN BLOQUEAR el archivo original"""
    ruta_original = os.path.join(RESPALDO_DIR, nombre)

    if os.path.exists(ruta_original):
        # Crear una copia temporal
        temp_dir = tempfile.gettempdir()
        copia_temporal = os.path.join(temp_dir, f"copia_{nombre}")
        shutil.copy2(ruta_original, copia_temporal)

        archivo = open(copia_temporal, 'rb')
        return FileResponse(archivo, as_attachment=True, filename=nombre)

    return HttpResponseNotFound("Archivo no encontrado.")


#--------------------------------------# Eliminar respaldo

import threading
import time

import threading
import time
import os

def eliminar_respaldo(request, nombre):
    """Elimina el respaldo con más intentos para asegurar eliminación en Windows"""
    ruta = os.path.join(RESPALDO_DIR, nombre)

    def intentar_eliminar(path):
        max_intentos = 15
        for i in range(max_intentos):
            try:
                os.remove(path)
                print(f"✅ Respaldo eliminado correctamente: {path}")
                break
            except PermissionError:
                print(f"⚠️ Intento {i+1}/{max_intentos}: El archivo está en uso. Esperando...")
                time.sleep(1.5)  # Esperar 1.5 segundos entre intentos
            except Exception as e:
                print(f"❌ Error inesperado: {e}")
                break
        else:
            print(f"❌ No se pudo eliminar el archivo tras {max_intentos} intentos: {path}")

    if os.path.exists(ruta):
        threading.Thread(target=intentar_eliminar, args=(ruta,)).start()

    return redirect('backup:listar_respaldos')


#--------------------------------------# Listar respaldos    
from django.http import HttpResponse
import traceback

def listar_respaldos(request):
    try:
        archivos = []
        if os.path.exists(RESPALDO_DIR):
            for archivo in os.listdir(RESPALDO_DIR):
                ruta = os.path.join(RESPALDO_DIR, archivo)
                archivos.append({
                    'nombre': archivo,
                    'tamano': f"{os.path.getsize(ruta) / 1024:.2f} KB",
                    'fecha': now().strftime('%d/%m/%Y %H:%M'),
                })

        ultima_fecha = archivos[-1]['fecha'] if archivos else 'No disponible'

        return render(request, 'backup/respaldo_db.html', {
            'respaldos': archivos,
            'ultima_fecha': ultima_fecha
        })
    except Exception as e:
        return HttpResponse(f"""
            <h2>⚠️ Error en listar_respaldos:</h2>
            <pre>{traceback.format_exc()}</pre>
        """, status=500)

#-------------Provedores a Excel----------------------------
import openpyxl
from django.http import HttpResponse
from apps.CarritoApp.models import Provedor
def exportar_provedores_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    # Encabezados
    ws.append(["ID", "Nombre", "Dirección", "Teléfono", "Email"])
    # Filas de datos desde la base
    for p in Provedor.objects.all():
        ws.append([p.id, p.nombre, p.direccion or '', p.telefono or '', p.email or ''])
    # Preparar respuesta para descargar
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=proveedores.xlsx'
    wb.save(response)
    return response

#-------------Categorias a Excel----------------------------
import openpyxl
from django.http import HttpResponse
from apps.CarritoApp.models import Categ_producto  # Asegurate de que esté bien la importación

def exportar_categorias_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categorías"

    # Encabezados
    ws.append(["ID", "Nombre"])

    # Datos desde la base
    for categoria in Categ_producto.objects.all():
        ws.append([categoria.id, categoria.nombre])

    # Preparar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=categorias.xlsx'
    wb.save(response)
    return response

#-------------Producto a Excel----------------------------
import openpyxl
from django.http import HttpResponse
from apps.CarritoApp.models import Producto

def exportar_productos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezado
    ws.append(["ID", "Nombre", "Descripción", "Categoría", "Stock", "Precio"])

    # Filas
    for producto in Producto.objects.select_related('categoria').all():
        ws.append([
            producto.id,
            producto.nombre_producto,
            producto.descripcion,
            producto.categoria.nombre if producto.categoria else '',
            producto.stock,
            float(producto.precio)
        ])

    # Generar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)
    return response


#-------------Compra a Excel----------------------------
import openpyxl
from django.http import HttpResponse
from apps.CarritoApp.models import Compra

def exportar_compras_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras"

    # Encabezado
    ws.append(["ID", "Producto", "Cantidad", "Precio Compra", "Factura", "Fecha", "Proveedor"])

    # Filas
    for compra in Compra.objects.select_related('producto', 'provedor'):
        ws.append([
            compra.id,
            compra.producto.nombre_producto if compra.producto else '',
            compra.cantidad,
            float(compra.precio_compra),
            compra.factura_compra,
            compra.fecha_compra.strftime('%d/%m/%Y'),
            compra.provedor.nombre if compra.provedor else '',
        ])

    # Generar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=compras.xlsx'
    wb.save(response)
    return response

#-------------Factura a Excel----------------------------

import openpyxl
from django.http import HttpResponse
from apps.CarritoApp.models import Factura

def exportar_facturas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"

    # Encabezados
    ws.append([
        "Número", "Fecha", "Cliente", "DNI", "CUIL", "IVA", "Domicilio",
        "Método de Pago", "Total", "Total con Interés", "Cuotas", "Cuota Mensual",
        "Tarjeta", "Ticket", "Estado Crédito", "Estado Entrega", "Vendedor"
    ])

    # Datos
    for f in Factura.objects.all():
        metodo_pago_nombre = f.metodo_pago.tarjeta_nombre if f.metodo_pago else f.metodo_pago_manual or ""

        ws.append([
            f.numero_factura,
            f.fecha.strftime('%d/%m/%Y'),
            f"{f.nombre_cliente} {f.apellido_cliente}",
            f.dni_cliente or "",
            f.cuil or "",
            f.iva or "",
            f.domicilio or "",
            metodo_pago_nombre,
            float(f.total),
            float(f.total_con_interes),
            f.cuotas,
            float(f.cuota_mensual),
            f.tarjeta_nombre or "",
            f.numero_tiket or "",
            f.estado_credito,
            f.estado_entrega,
            f.vendedor
        ])

    # Generar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=facturas.xlsx'
    wb.save(response)
    return response

#-------------Cuenta corriente a Excel----------------------------
import openpyxl
from django.http import HttpResponse
from apps.CarritoApp.models import CuentaCorriente  # Ajustá según nombre real de tu app

def exportar_cuenta_corriente_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuenta Corriente"

    # Encabezados
    ws.append([
        "Número de Factura", "Descripción", "Fecha Cuota", "Total con Interés",
        "Cuotas Totales", "Cuotas Pagadas", "Cuotas Restantes", "Suma de Cuotas",
        "Importe Mensual", "Importe Pagado", "Entrega a Cuenta", "Interés Aplicado",
        "Método de Pago", "Tarjeta", "Número Tarjeta", "Estado del Crédito"
    ])

    # Datos
    for cc in CuentaCorriente.objects.all():
        metodo_pago = cc.metodo_pago.tarjeta_nombre if cc.metodo_pago else "Efectivo"
        estado_credito = cc.estado_credito
        tarjeta_numero = f"****{cc.tarjeta_numero[-4:]}" if cc.tarjeta_numero else ""

        ws.append([
            cc.numero_factura,
            cc.descripcion,
            cc.fecha_cuota.strftime('%d/%m/%Y'),
            float(cc.total_con_interes),
            cc.cuota_total,
            cc.cuota_paga,
            cc.cuota_debe,
            cc.cuota_suma,
            float(cc.imp_mensual),
            float(cc.imp_cuota_pagadas),
            float(cc.entrega_cta),
            float(cc.interes_aplicado),
            metodo_pago,
            cc.tarjeta_nombre or "",
            tarjeta_numero,
            estado_credito,
        ])

    # Generar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cuenta_corriente.xlsx'
    wb.save(response)
    return response
